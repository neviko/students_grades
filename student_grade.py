
from openpyxl import load_workbook
from students_grades.obj.student import Student
from students_grades.obj.course import Course


gradesFilePath = "C:\College Programing\Python\Final project\students_grades\grades.XLSX"
exemptionsFilePath = "C:\College Programing\Python\Final project\students_grades\exemption.XLSX"
lectureSheetPath = "C:\College Programing\Python\Final project\students_grades\lectureSheet.XLSX"

weight_final_project=0.0
numOfRequiredEx=0


def generate_students_obj(grades_filename):
    student_obj_list = []
    wb = load_workbook(grades_filename, read_only=True)
    ws = wb['FileFromYedion']
    rowCounter=0

    global weight_final_project
    weight_final_project= ws["B3"].value

    global numOfRequiredEx
    numOfRequiredEx = ws["C3"].value

    Course.weight_final_project = weight_final_project
    Course.numOfRequiredEx = numOfRequiredEx


    for row in ws.rows:

        rowCounter +=1
        # Ignore the first two lines
        if rowCounter < 9 :
            continue
        # if row[0].row < 3:
        #     continue

        # Create list of student data [id, lastName, name, grade_1, grade_2]
        curr_student_list = []
        for cell in row:
            if cell.value == "":
                curr_student_list.append(0)
            else:
                curr_student_list.append(cell.value)

        # The excel file is from right to left -> reverse
        reversed(curr_student_list)
        curr_student_obj = Student(curr_student_list,numOfRequiredEx,weight_final_project)
        student_obj_list.append(curr_student_obj)

    wb._archive.close() # close streaming
    return student_obj_list


def generate_exemption_dict(exemp_filename):
    exemption_dict = {}
    wb = load_workbook(exemp_filename, read_only=True)
    ws = wb['Sheet1']

    for row in ws.rows:
        # Ignore the first line
        if row[0].row == 1 :
            continue

        id = row[0].value
        assert type(id) is int

        exemp = row[2].value
        exemption_dict[id] = exemp

    return exemption_dict



def generate_finalGrades(student_obj_list):

    for student in student_obj_list:

        # get exercise average
        exGradesList = sorted(student.ex_grade_list, key=int, reverse=True)  # descending order
        highestGrades = exGradesList[0:student.numOfRequiredEx]  # highest required grades
        exAvg = sum(highestGrades) / float(len(highestGrades))

        finalGrade = (1 - course.weight_final_project) * exAvg + course.weight_final_project * student.ex_grade_list[-1]

        student.final_grade = finalGrade


def gradesToLectureSheet(student_obj_list):

    wb = load_workbook(lectureSheetPath)
    ws = wb['1']
    rowCounter = 0

    for row in ws.rows:

        rowCounter +=1
        # Ignore the first two lines
        if rowCounter < 15 :
            continue

        if row[0] in student_obj_list:
            studentGrade =


if __name__ == '__main__':
    course_name = 'עקרונות התכנות בפייתון'
    course = Course(course_name, numOfRequiredEx, weight_final_project)
    course.student_obj_list = generate_students_obj(gradesFilePath)
    course.weight_final_project = weight_final_project
    course.num_ex_required = weight_final_project
    course.set_exemptions(generate_exemption_dict(exemptionsFilePath))
    generate_finalGrades(course.student_obj_list)
    course.print_details()

